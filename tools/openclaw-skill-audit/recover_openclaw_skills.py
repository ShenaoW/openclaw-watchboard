#!/usr/bin/env python3
"""Recover historical skills from openclaw/skills with a lightweight git mirror."""

from __future__ import annotations

import argparse
import csv
import io
import json
import shutil
import subprocess
import tarfile
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

DEFAULT_REMOTE = "https://github.com/openclaw/skills.git"


@dataclass(frozen=True)
class SkillSlug:
    author: str
    skill: str

    @property
    def slug(self) -> str:
        return f"{self.author}/{self.skill}"

    @property
    def repo_path(self) -> str:
        return f"skills/{self.author}/{self.skill}"


def run_git(repo: Path, args: list[str], check: bool = True) -> subprocess.CompletedProcess:
    cmd = ["git", "-C", str(repo), *args]
    return subprocess.run(cmd, check=check, text=True, capture_output=True)


def parse_skill_from_path(path: str) -> SkillSlug | None:
    parts = path.strip().split("/")
    if len(parts) < 3 or parts[0] != "skills":
        return None
    author, skill = parts[1], parts[2]
    if not author or not skill:
        return None
    return SkillSlug(author=author, skill=skill)


def iter_paths_from_name_status(line: str) -> Iterable[str]:
    line = line.strip("\n")
    if not line:
        return []
    parts = line.split("\t")
    if len(parts) < 2:
        return []
    # Format examples:
    # A\tpath
    # M\tpath
    # R100\told\tnew
    return parts[1:]


def parse_name_status_paths(line: str) -> list[tuple[str, bool]]:
    """Parse `git log --name-status` line.

    Returns tuples of (path, exists_after_commit).
    """
    line = line.strip("\n")
    if not line:
        return []
    parts = line.split("\t")
    if len(parts) < 2:
        return []
    status = parts[0].strip()
    paths = parts[1:]

    # Rename: old path removed, new path exists.
    if status.startswith("R") and len(paths) >= 2:
        return [(paths[0], False), (paths[1], True)]
    # Copy: both source and destination exist.
    if status.startswith("C") and len(paths) >= 2:
        return [(paths[0], True), (paths[1], True)]
    # Delete: path no longer exists after commit.
    exists = not status.startswith("D")
    return [(p, exists) for p in paths]


def init_repo(repo: Path, remote: str, ref: str, clone_mode: str) -> None:
    if repo.exists() and (repo / ".git").exists():
        run_git(repo, ["fetch", "--all", "--tags", "--prune"])
        run_git(repo, ["checkout", ref])
        run_git(repo, ["pull", "--ff-only"])
        return

    repo.parent.mkdir(parents=True, exist_ok=True)
    cmd = ["git", "clone", "--single-branch", "--branch", ref]
    if clone_mode == "partial":
        cmd.extend(["--filter=blob:none", "--no-checkout"])
    cmd.extend([remote, str(repo)])
    subprocess.run(cmd, check=True, text=True)


def collect_history_slugs(repo: Path) -> set[SkillSlug]:
    cp = run_git(repo, ["log", "--name-status", "--pretty=format:", "--", "skills"])
    slugs: set[SkillSlug] = set()
    for line in cp.stdout.splitlines():
        for path, _ in parse_name_status_paths(line):
            slug = parse_skill_from_path(path)
            if slug:
                slugs.add(slug)
    return slugs


def collect_current_slugs(repo: Path, ref: str) -> set[SkillSlug]:
    cp = run_git(repo, ["ls-tree", "-r", "--name-only", ref, "--", "skills"])
    slugs: set[SkillSlug] = set()
    for line in cp.stdout.splitlines():
        slug = parse_skill_from_path(line)
        if slug:
            slugs.add(slug)
    return slugs


def collect_latest_commit_maps(repo: Path) -> tuple[dict[str, str], dict[str, str]]:
    """Collect per-skill latest commit info in one history scan.

    latest_touch_commit: latest commit where slug appeared in diff paths.
    latest_recover_commit: latest commit where slug exists after the commit.
    """
    latest_touch: dict[str, str] = {}
    latest_recover: dict[str, str] = {}
    cmd = ["git", "-C", str(repo), "log", "--name-status", "--pretty=format:@@@%H", "--", "skills"]
    with subprocess.Popen(cmd, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE) as proc:
        current_commit = ""
        assert proc.stdout is not None
        for line in proc.stdout:
            line = line.rstrip("\n")
            if line.startswith("@@@"):
                current_commit = line[3:].strip()
                continue
            if not line or not current_commit:
                continue
            for path, exists_after in parse_name_status_paths(line):
                slug = parse_skill_from_path(path)
                if not slug:
                    continue
                key = slug.slug.lower()
                if key not in latest_touch:
                    latest_touch[key] = current_commit
                if exists_after and key not in latest_recover:
                    latest_recover[key] = current_commit
    return latest_touch, latest_recover


def write_inventory(repo: Path, ref: str, out_csv: Path, out_json: Path | None) -> dict[str, object]:
    history_slugs = collect_history_slugs(repo)
    current_slugs = collect_current_slugs(repo, ref)
    latest_touch_map, latest_recover_map = collect_latest_commit_maps(repo)
    merged = sorted(history_slugs | current_slugs, key=lambda x: (x.author, x.skill))

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "author",
                "skill",
                "slug",
                "current",
                "seen_in_history",
                "latest_touch_commit",
                "latest_recover_commit",
            ],
        )
        writer.writeheader()
        for slug in merged:
            slug_key = slug.slug.lower()
            writer.writerow(
                {
                    "author": slug.author,
                    "skill": slug.skill,
                    "slug": slug.slug,
                    "current": "1" if slug in current_slugs else "0",
                    "seen_in_history": "1" if slug in history_slugs else "0",
                    "latest_touch_commit": latest_touch_map.get(slug_key, ""),
                    "latest_recover_commit": latest_recover_map.get(slug_key, ""),
                }
            )

    payload = {
        "total_skills": len(merged),
        "current_skills": sum(1 for s in merged if s in current_slugs),
        "deleted_or_missing_from_head": sum(1 for s in merged if s not in current_slugs),
        "inventory_csv": str(out_csv),
    }
    if out_json:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def read_inventory(csv_path: Path) -> list[SkillSlug]:
    slugs: list[SkillSlug] = []
    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            slug = row.get("slug", "").strip()
            if "/" not in slug:
                continue
            author, skill = slug.split("/", 1)
            if author and skill:
                slugs.append(SkillSlug(author=author, skill=skill))
    return slugs


def read_inventory_rows(csv_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def build_index(
    inventory_rows: list[dict[str, str]],
) -> tuple[dict[str, dict[str, str]], dict[str, set[str]], dict[str, set[str]]]:
    by_slug: dict[str, dict[str, str]] = {}
    by_author: dict[str, set[str]] = {}
    by_skill: dict[str, set[str]] = {}
    for row in inventory_rows:
        slug = (row.get("slug") or "").strip().lower()
        if not slug or "/" not in slug:
            continue
        by_slug[slug] = row
        author, skill = slug.split("/", 1)
        by_author.setdefault(author, set()).add(slug)
        by_skill.setdefault(skill, set()).add(slug)
    return by_slug, by_author, by_skill


def normalize_id(value: str) -> str:
    return value.strip().lower()


def has_meaningful(value: str) -> bool:
    v = normalize_id(value)
    return v not in {"", "unknown", "n/a", "na", "none", "null", "-"}


def pick_first_meaningful(values: list[str]) -> str:
    for v in values:
        if has_meaningful(v):
            return v.strip()
    return ""


def merge_values(values: list[str]) -> str:
    items: list[str] = []
    seen: set[str] = set()
    for raw in values:
        v = str(raw or "").strip()
        if not v:
            continue
        key = v.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append(v)
    return " | ".join(items)


def resolve_row_slugs(
    author_name: str, skill_name: str, by_slug: dict[str, dict[str, str]], by_author: dict[str, set[str]], by_skill: dict[str, set[str]]
) -> list[str]:
    author = normalize_id(author_name)
    skill = normalize_id(skill_name)

    if author and skill:
        s = f"{author}/{skill}"
        return [s] if s in by_slug else []
    if author:
        return sorted(by_author.get(author, set()))
    if skill:
        return sorted(by_skill.get(skill, set()))
    return []


def resolve_targets(targets_file: Path, inventory: list[SkillSlug]) -> tuple[set[SkillSlug], list[str]]:
    by_author: dict[str, set[SkillSlug]] = {}
    by_skill: dict[str, set[SkillSlug]] = {}
    by_slug: dict[str, SkillSlug] = {}
    for s in inventory:
        by_slug[s.slug.lower()] = s
        by_author.setdefault(s.author.lower(), set()).add(s)
        by_skill.setdefault(s.skill.lower(), set()).add(s)

    selected: set[SkillSlug] = set()
    unmatched: list[str] = []
    for raw in targets_file.read_text(encoding="utf-8").splitlines():
        item = raw.strip()
        if not item or item.startswith("#"):
            continue
        mode = "auto"
        value = item
        if ":" in item:
            maybe_mode, maybe_value = item.split(":", 1)
            maybe_mode = maybe_mode.strip().lower()
            if maybe_mode in {"author", "skill", "slug"}:
                mode = maybe_mode
                value = maybe_value.strip()

        value_lc = value.lower()
        if mode == "slug" or "/" in value:
            v = by_slug.get(value_lc)
            if v:
                selected.add(v)
            else:
                unmatched.append(item)
            continue

        if mode == "author":
            group = by_author.get(value_lc, set())
            if group:
                selected.update(group)
            else:
                unmatched.append(item)
            continue

        if mode == "skill":
            group = by_skill.get(value_lc, set())
            if group:
                selected.update(group)
            else:
                unmatched.append(item)
            continue

        # auto mode (default): try author first, then skill
        group = by_author.get(value_lc, set())
        if group:
            selected.update(group)
            continue
        group = by_skill.get(value_lc, set())
        if group:
            selected.update(group)
            continue
        unmatched.append(item)
    return selected, unmatched


def skill_exists_at_commit(repo: Path, commit: str, slug: SkillSlug) -> bool:
    cp = run_git(repo, ["cat-file", "-e", f"{commit}:{slug.repo_path}"], check=False)
    return cp.returncode == 0


def get_tree_sha(repo: Path, commit: str, slug: SkillSlug) -> str | None:
    cp = run_git(repo, ["rev-parse", f"{commit}:{slug.repo_path}"], check=False)
    if cp.returncode != 0:
        return None
    return cp.stdout.strip()


def list_path_commits(repo: Path, ref: str, slug: SkillSlug) -> list[str]:
    cp = run_git(repo, ["rev-list", ref, "--", slug.repo_path])
    commits = [c for c in cp.stdout.splitlines() if c.strip()]
    return commits


def find_latest_existing_commit(repo: Path, ref: str, slug: SkillSlug) -> str | None:
    for commit in list_path_commits(repo, ref, slug):
        if skill_exists_at_commit(repo, commit, slug):
            return commit
    return None


def resolve_latest_existing_commits(repo: Path, target_slugs: set[str]) -> dict[str, str]:
    if not target_slugs:
        return {}
    remaining = set(target_slugs)
    found: dict[str, str] = {}
    cmd = ["git", "-C", str(repo), "log", "--name-status", "--pretty=format:@@@%H", "--", "skills"]
    with subprocess.Popen(cmd, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE) as proc:
        current_commit = ""
        assert proc.stdout is not None
        for line in proc.stdout:
            line = line.rstrip("\n")
            if line.startswith("@@@"):
                current_commit = line[3:].strip()
                continue
            if not line or not current_commit:
                continue
            parts = line.split("\t")
            if len(parts) < 2:
                continue
            status = parts[0].strip()
            if status.startswith("D"):
                continue
            for path in parts[1:]:
                skill = parse_skill_from_path(path)
                if not skill:
                    continue
                slug = skill.slug.lower()
                if slug in remaining:
                    found[slug] = current_commit
                    remaining.remove(slug)
            if not remaining:
                proc.terminate()
                break
        proc.wait()
    return found


def export_skill_snapshot(
    repo: Path, slug: SkillSlug, commit: str, dst_dir: Path, force: bool = False
) -> None:
    if dst_dir.exists():
        if not force:
            raise FileExistsError(f"Destination already exists: {dst_dir}")
        shutil.rmtree(dst_dir)
    dst_dir.parent.mkdir(parents=True, exist_ok=True)

    raw = subprocess.run(
        ["git", "-C", str(repo), "archive", "--format=tar", commit, slug.repo_path],
        check=True,
        capture_output=True,
    ).stdout
    with tempfile.TemporaryDirectory(prefix="openclaw-skill-") as tmpdir:
        tmp = Path(tmpdir)
        with tarfile.open(fileobj=io.BytesIO(raw), mode="r:") as tf:
            tf.extractall(tmp)
        src_dir = tmp / slug.repo_path
        if not src_dir.exists():
            raise FileNotFoundError(f"Skill path missing from archive: {slug.repo_path}")
        shutil.copytree(src_dir, dst_dir)


def export_targets(
    repo: Path,
    ref: str,
    targets_file: Path,
    inventory_csv: Path,
    out_dir: Path,
    mode: str,
    force: bool,
) -> dict[str, object]:
    inventory = read_inventory(inventory_csv)
    selected, unmatched = resolve_targets(targets_file, inventory)

    report: dict[str, object] = {
        "targets_file": str(targets_file),
        "selected_count": len(selected),
        "unmatched": unmatched,
        "exported": [],
        "skipped": [],
    }

    for slug in sorted(selected, key=lambda s: (s.author, s.skill)):
        if mode == "latest":
            commit = find_latest_existing_commit(repo, ref, slug)
            if not commit:
                report["skipped"].append({"slug": slug.slug, "reason": "not recoverable"})
                continue
            dst = out_dir / slug.author / slug.skill / f"latest_{commit[:12]}"
            export_skill_snapshot(repo, slug, commit, dst, force=force)
            report["exported"].append({"slug": slug.slug, "commit": commit, "path": str(dst)})
            continue

        commits = list_path_commits(repo, ref, slug)
        seen_trees: set[str] = set()
        version_idx = 0
        for commit in commits:
            tree = get_tree_sha(repo, commit, slug)
            if not tree:
                continue
            if tree in seen_trees:
                continue
            seen_trees.add(tree)
            version_idx += 1
            dst = out_dir / slug.author / slug.skill / "versions" / f"{version_idx:04d}_{commit[:12]}"
            export_skill_snapshot(repo, slug, commit, dst, force=force)
            report["exported"].append({"slug": slug.slug, "commit": commit, "path": str(dst)})
        if version_idx == 0:
            report["skipped"].append({"slug": slug.slug, "reason": "not recoverable"})

    return report


def export_from_xlsx(
    repo: Path,
    ref: str,
    inventory_csv: Path,
    input_xlsx: Path,
    output_xlsx: Path,
    out_dir: Path,
    mode: str,
    force: bool,
    skip_extract: bool,
    deleted_only: bool,
    report_json: Path | None,
) -> dict[str, object]:
    from openpyxl import Workbook, load_workbook

    inventory_rows = read_inventory_rows(inventory_csv)
    by_slug, by_author, by_skill = build_index(inventory_rows)

    wb = load_workbook(input_xlsx, data_only=True)
    output_rows: list[dict[str, str]] = []
    source_extra_cols: list[str] = []
    report: dict[str, object] = {
        "input_xlsx": str(input_xlsx),
        "output_xlsx": str(output_xlsx),
        "rows_total": 0,
        "rows_with_match": 0,
        "rows_extracted_success": 0,
        "source_sheets_processed": [],
        "skipped_sheets": [],
        "unique_matched_slugs": 0,
        "deleted_only": deleted_only,
        "exported_items": [],
        "failed_rows": [],
    }

    matched_entries: list[dict[str, object]] = []

    for ws in wb.worksheets:
        raw_headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        header_pos: dict[str, list[int]] = {}
        for idx, name in enumerate(raw_headers, start=1):
            key = name.strip()
            if not key:
                continue
            header_pos.setdefault(key, []).append(idx)

        has_author_col = "author_name" in header_pos
        has_skill_col = "skill_name" in header_pos
        if not has_author_col and not has_skill_col:
            report["skipped_sheets"].append({"sheet": ws.title, "reason": "no author_name/skill_name columns"})
            continue
        report["source_sheets_processed"].append(ws.title)

        row_meta_cols = [c for c in header_pos.keys() if c not in {"author_name", "skill_name"}]
        for col in row_meta_cols:
            if col not in source_extra_cols:
                source_extra_cols.append(col)

        for row in range(2, ws.max_row + 1):
            report["rows_total"] += 1
            author_values = [str(ws.cell(row=row, column=i).value or "").strip() for i in header_pos.get("author_name", [])]
            skill_values = [str(ws.cell(row=row, column=i).value or "").strip() for i in header_pos.get("skill_name", [])]
            author_query = pick_first_meaningful(author_values)
            skill_query = pick_first_meaningful(skill_values)

            if not author_query and not skill_query:
                report["failed_rows"].append({"sheet": ws.title, "row": row, "reason": "both author_name and skill_name empty/unknown"})
                continue

            slugs = resolve_row_slugs(author_query, skill_query, by_slug, by_author, by_skill)
            if deleted_only:
                slugs = [s for s in slugs if (by_slug.get(s, {}).get("current") or "") != "1"]

            if not slugs:
                report["failed_rows"].append(
                    {
                        "sheet": ws.title,
                        "row": row,
                        "author_name": author_query,
                        "skill_name": skill_query,
                        "reason": "no match" if not deleted_only else "no deleted match",
                    }
                )
                continue
            report["rows_with_match"] += 1

            meta_map = {col: merge_values([ws.cell(row=row, column=i).value for i in header_pos.get(col, [])]) for col in row_meta_cols}
            for slug in slugs:
                matched_entries.append(
                    {
                        "sheet": ws.title,
                        "row": row,
                        "slug": slug,
                        "meta": meta_map,
                    }
                )

    unique_slugs = {str(e["slug"]) for e in matched_entries}
    report["unique_matched_slugs"] = len(unique_slugs)

    latest_commit_map: dict[str, str] = {
        s: (by_slug.get(s, {}).get("latest_recover_commit") or "").strip() for s in unique_slugs
    }
    if mode == "latest":
        missing = {s for s, c in latest_commit_map.items() if not c}
        if missing:
            latest_commit_map.update(resolve_latest_existing_commits(repo, missing))

    exported_once: set[tuple[str, str]] = set()
    slug_groups: dict[str, dict[str, object]] = {}
    for entry in matched_entries:
        slug = str(entry["slug"])
        sheet = str(entry["sheet"])
        row = int(entry["row"])
        meta_map = dict(entry["meta"])
        group = slug_groups.setdefault(
            slug,
            {
                "source_sheets": [],
                "source_refs": [],
                "meta_values": {col: [] for col in source_extra_cols},
            },
        )
        source_sheets = group["source_sheets"]
        assert isinstance(source_sheets, list)
        if sheet not in source_sheets:
            source_sheets.append(sheet)

        source_refs = group["source_refs"]
        assert isinstance(source_refs, list)
        ref = f"{sheet}:{row}"
        if ref not in source_refs:
            source_refs.append(ref)

        meta_values = group["meta_values"]
        assert isinstance(meta_values, dict)
        for col in source_extra_cols:
            v = str(meta_map.get(col, "")).strip()
            if not v:
                continue
            bucket = meta_values.setdefault(col, [])
            if v not in bucket:
                bucket.append(v)

    for slug in sorted(slug_groups.keys()):
        group = slug_groups[slug]
        author, skill = slug.split("/", 1)
        current_exists = (by_slug.get(slug, {}).get("current") or "") == "1"
        repo_deleted = not current_exists

        extraction_success = False
        latest_commit = ""
        if mode == "latest":
            latest_commit = latest_commit_map.get(slug, "")
            if latest_commit:
                if skip_extract:
                    extraction_success = True
                else:
                    cache_key = (slug, latest_commit)
                    try:
                        if cache_key not in exported_once:
                            dst = out_dir / author / skill / f"latest_{latest_commit[:12]}"
                            export_skill_snapshot(repo, SkillSlug(author, skill), latest_commit, dst, force=force)
                            exported_once.add(cache_key)
                            report["exported_items"].append(
                                {
                                    "sheet": merge_values(group["source_sheets"]),
                                    "row": merge_values(group["source_refs"]),
                                    "slug": slug,
                                    "commit": latest_commit,
                                    "path": str(dst),
                                }
                            )
                        extraction_success = True
                    except Exception as exc:
                        report["failed_rows"].append(
                            {
                                "sheet": merge_values(group["source_sheets"]),
                                "row": merge_values(group["source_refs"]),
                                "slug": slug,
                                "reason": f"export_failed: {exc}",
                            }
                        )
        else:
            # all-mode keeps legacy behavior (slower)
            if skip_extract:
                extraction_success = bool(list_path_commits(repo, ref, SkillSlug(author, skill)))
            else:
                commits = list_path_commits(repo, ref, SkillSlug(author, skill))
                seen_trees: set[str] = set()
                version_idx = 0
                for commit in commits:
                    tree = get_tree_sha(repo, commit, SkillSlug(author, skill))
                    if not tree or tree in seen_trees:
                        continue
                    seen_trees.add(tree)
                    version_idx += 1
                    cache_key = (slug, commit)
                    if cache_key not in exported_once:
                        dst = out_dir / author / skill / "versions" / f"{version_idx:04d}_{commit[:12]}"
                        export_skill_snapshot(repo, SkillSlug(author, skill), commit, dst, force=force)
                        exported_once.add(cache_key)
                        report["exported_items"].append(
                            {"sheet": sheet, "row": row, "slug": slug, "commit": commit, "path": str(dst)}
                        )
                extraction_success = version_idx > 0

        if extraction_success:
            report["rows_extracted_success"] += 1

        merged_meta: dict[str, str] = {}
        meta_values = group["meta_values"]
        assert isinstance(meta_values, dict)
        for col in source_extra_cols:
            vals = meta_values.get(col, [])
            if not isinstance(vals, list):
                vals = []
            merged_meta[col] = merge_values(vals)

        out_row: dict[str, str] = {
            "author_name": author,
            "skill_name": skill,
            "slug": slug,
            "source_sheet": merge_values(group["source_sheets"]),
            "source_row": merge_values(group["source_refs"]),
            "latest_recover_commit": latest_commit,
            "仓库是否删除": "true" if repo_deleted else "false",
            "当前是否存在": "true" if current_exists else "false",
            "提取是否成功": "true" if extraction_success else "false",
        }
        out_row.update(merged_meta)
        output_rows.append(out_row)

    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = "RecoveredSkills"
    final_headers = [
        "author_name",
        "skill_name",
        "slug",
        "source_sheet",
        "source_row",
        "latest_recover_commit",
        *source_extra_cols,
        "仓库是否删除",
        "当前是否存在",
        "提取是否成功",
    ]
    result_ws.append(final_headers)
    for row in output_rows:
        result_ws.append([row.get(h, "") for h in final_headers])

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    result_wb.save(output_xlsx)

    if report_json:
        report_json.parent.mkdir(parents=True, exist_ok=True)
        report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def export_deleted_all(
    repo: Path,
    ref: str,
    inventory_csv: Path,
    output_xlsx: Path,
    out_dir: Path,
    mode: str,
    force: bool,
    skip_extract: bool,
    report_json: Path | None,
) -> dict[str, object]:
    from openpyxl import Workbook

    inventory_rows = read_inventory_rows(inventory_csv)
    deleted_rows = [r for r in inventory_rows if (r.get("current") or "").strip() != "1"]
    deleted_rows.sort(key=lambda r: ((r.get("author") or ""), (r.get("skill") or "")))

    deleted_slugs = {
        (r.get("slug") or "").strip().lower() for r in deleted_rows if (r.get("slug") or "").strip()
    }
    deleted_slugs.discard("")

    latest_commit_map: dict[str, str] = {}
    if mode == "latest":
        for r in deleted_rows:
            slug = (r.get("slug") or "").strip().lower()
            if not slug:
                continue
            latest_commit_map[slug] = (r.get("latest_recover_commit") or "").strip()
        missing = {s for s in deleted_slugs if not latest_commit_map.get(s)}
        if missing:
            latest_commit_map.update(resolve_latest_existing_commits(repo, missing))

    result_wb = Workbook()
    ws = result_wb.active
    ws.title = "DeletedSkills"
    headers = [
        "author_name",
        "skill_name",
        "slug",
        "latest_recover_commit",
        "latest_touch_commit",
        "仓库是否删除",
        "当前是否存在",
        "提取是否成功",
    ]
    ws.append(headers)

    report: dict[str, object] = {
        "inventory_csv": str(inventory_csv),
        "output_xlsx": str(output_xlsx),
        "deleted_slug_count": len(deleted_slugs),
        "rows_written": 0,
        "rows_extracted_success": 0,
        "exported_items": [],
        "failed_slugs": [],
    }

    exported_once: set[tuple[str, str]] = set()
    for row in deleted_rows:
        slug = (row.get("slug") or "").strip().lower()
        if not slug or "/" not in slug:
            continue
        author, skill = slug.split("/", 1)
        latest_touch = (row.get("latest_touch_commit") or "").strip()
        latest_recover = latest_commit_map.get(slug, "")
        extraction_success = False

        if mode == "latest":
            if latest_recover:
                if skip_extract:
                    extraction_success = True
                else:
                    cache_key = (slug, latest_recover)
                    try:
                        if cache_key not in exported_once:
                            dst = out_dir / author / skill / f"latest_{latest_recover[:12]}"
                            export_skill_snapshot(repo, SkillSlug(author, skill), latest_recover, dst, force=force)
                            exported_once.add(cache_key)
                            report["exported_items"].append(
                                {"slug": slug, "commit": latest_recover, "path": str(dst)}
                            )
                        extraction_success = True
                    except Exception as exc:
                        report["failed_slugs"].append({"slug": slug, "reason": f"export_failed: {exc}"})
            else:
                report["failed_slugs"].append({"slug": slug, "reason": "missing latest_recover_commit"})
        else:
            commits = list_path_commits(repo, ref, SkillSlug(author, skill))
            seen_trees: set[str] = set()
            version_idx = 0
            try:
                for commit in commits:
                    tree = get_tree_sha(repo, commit, SkillSlug(author, skill))
                    if not tree or tree in seen_trees:
                        continue
                    seen_trees.add(tree)
                    version_idx += 1
                    if skip_extract:
                        continue
                    cache_key = (slug, commit)
                    if cache_key not in exported_once:
                        dst = out_dir / author / skill / "versions" / f"{version_idx:04d}_{commit[:12]}"
                        export_skill_snapshot(repo, SkillSlug(author, skill), commit, dst, force=force)
                        exported_once.add(cache_key)
                        report["exported_items"].append(
                            {"slug": slug, "commit": commit, "path": str(dst)}
                        )
                extraction_success = version_idx > 0
            except Exception as exc:
                report["failed_slugs"].append({"slug": slug, "reason": f"export_failed: {exc}"})

        ws.append(
            [
                author,
                skill,
                slug,
                latest_recover,
                latest_touch,
                "true",
                "false",
                "true" if extraction_success else "false",
            ]
        )
        report["rows_written"] += 1
        if extraction_success:
            report["rows_extracted_success"] += 1

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    result_wb.save(output_xlsx)
    if report_json:
        report_json.parent.mkdir(parents=True, exist_ok=True)
        report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def export_full_clawhub_dataset(
    repo: Path,
    ref: str,
    out_root: Path,
    force: bool,
    skip_extract: bool,
) -> dict[str, object]:
    out_root.mkdir(parents=True, exist_ok=True)

    inventory_csv = out_root / "inventory_all_skills.csv"
    inventory_summary_json = out_root / "inventory_summary.json"
    manifest_csv = out_root / "clawhub_full_skills_manifest.csv"
    manifest_json = out_root / "clawhub_full_skills_manifest.json"
    skills_root = out_root / "skills"

    inventory_summary = write_inventory(
        repo=repo,
        ref=ref,
        out_csv=inventory_csv,
        out_json=inventory_summary_json,
    )
    inventory_rows = read_inventory_rows(inventory_csv)

    target_slugs = {
        (row.get("slug") or "").strip().lower()
        for row in inventory_rows
        if (row.get("slug") or "").strip()
    }
    latest_commit_map: dict[str, str] = {
        slug: (row.get("latest_recover_commit") or "").strip()
        for row in inventory_rows
        for slug in [((row.get("slug") or "").strip().lower())]
        if slug
    }
    missing = {slug for slug in target_slugs if not latest_commit_map.get(slug)}
    if missing:
        latest_commit_map.update(resolve_latest_existing_commits(repo, missing))

    exported_count = 0
    failed_count = 0
    manifest_rows: list[dict[str, str]] = []

    for row in sorted(inventory_rows, key=lambda r: ((r.get("author") or ""), (r.get("skill") or ""))):
        author = (row.get("author") or "").strip()
        skill = (row.get("skill") or "").strip()
        slug = (row.get("slug") or "").strip()
        slug_key = slug.lower()
        latest_recover_commit = latest_commit_map.get(slug_key, "")
        current = (row.get("current") or "").strip()
        seen_in_history = (row.get("seen_in_history") or "").strip()
        latest_touch_commit = (row.get("latest_touch_commit") or "").strip()

        export_path = ""
        export_status = "skipped"
        export_error = ""

        if latest_recover_commit:
            export_dir = skills_root / author / skill / f"latest_{latest_recover_commit[:12]}"
            export_path = str(export_dir)
            if skip_extract:
                export_status = "ready"
                exported_count += 1
            else:
                try:
                    export_skill_snapshot(
                        repo,
                        SkillSlug(author=author, skill=skill),
                        latest_recover_commit,
                        export_dir,
                        force=force,
                    )
                    export_status = "exported"
                    exported_count += 1
                except Exception as exc:
                    export_status = "failed"
                    export_error = str(exc)
                    failed_count += 1
        else:
            export_status = "missing_latest_recover_commit"
            failed_count += 1

        manifest_rows.append(
            {
                "author": author,
                "skill": skill,
                "slug": slug,
                "current": current,
                "seen_in_history": seen_in_history,
                "latest_touch_commit": latest_touch_commit,
                "latest_recover_commit": latest_recover_commit,
                "export_status": export_status,
                "export_path": export_path,
                "export_error": export_error,
            }
        )

    with manifest_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "author",
                "skill",
                "slug",
                "current",
                "seen_in_history",
                "latest_touch_commit",
                "latest_recover_commit",
                "export_status",
                "export_path",
                "export_error",
            ],
        )
        writer.writeheader()
        writer.writerows(manifest_rows)

    payload = {
        "out_root": str(out_root),
        "inventory_summary": inventory_summary,
        "inventory_csv": str(inventory_csv),
        "inventory_summary_json": str(inventory_summary_json),
        "manifest_csv": str(manifest_csv),
        "manifest_json": str(manifest_json),
        "skills_root": str(skills_root),
        "total_unique_skills": len(manifest_rows),
        "exported_count": exported_count,
        "failed_count": failed_count,
        "skip_extract": skip_extract,
    }
    manifest_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Recover all historical skills from openclaw/skills.")
    p.add_argument("--repo", type=Path, default=Path("repo/openclaw-skills"), help="Local git mirror path")
    p.add_argument("--ref", default="main", help="Git ref to scan/export from")
    p.add_argument("--remote", default=DEFAULT_REMOTE, help="Remote git URL")

    sub = p.add_subparsers(dest="command", required=True)

    init = sub.add_parser("init-repo", help="Clone/update openclaw/skills")
    init.add_argument(
        "--clone-mode",
        choices=["full", "partial"],
        default="full",
        help="Clone strategy: full checkout or partial no-checkout mirror",
    )

    inv = sub.add_parser("build-inventory", help="Build full skill inventory (including deleted)")
    inv.add_argument("--out-csv", type=Path, default=Path("output/inventory_all_skills.csv"))
    inv.add_argument("--out-json", type=Path, default=Path("output/inventory_summary.json"))

    exp = sub.add_parser("export", help="Export target skills by author/skill or author")
    exp.add_argument(
        "--targets",
        type=Path,
        required=True,
        help="Text file lines: author/skill, author, or skill (supports author:/skill:/slug: prefixes)",
    )
    exp.add_argument("--inventory-csv", type=Path, default=Path("output/inventory_all_skills.csv"))
    exp.add_argument("--out-dir", type=Path, default=Path("output/recovered_skills"))
    exp.add_argument("--mode", choices=["latest", "all"], default="latest")
    exp.add_argument("--force", action="store_true", help="Overwrite existing exported folders")
    exp.add_argument("--report-json", type=Path, default=Path("output/export_report.json"))

    xlsx = sub.add_parser("export-from-xlsx", help="Export by rows across all sheets and build a new status table")
    xlsx.add_argument("--input-xlsx", type=Path, required=True)
    xlsx.add_argument("--output-xlsx", type=Path, required=True)
    xlsx.add_argument("--inventory-csv", type=Path, default=Path("output/inventory_all_skills.csv"))
    xlsx.add_argument("--out-dir", type=Path, default=Path("output/recovered_skills_from_xlsx"))
    xlsx.add_argument("--mode", choices=["latest", "all"], default="latest")
    xlsx.add_argument("--force", action="store_true", help="Overwrite existing exported folders")
    xlsx.add_argument("--skip-extract", action="store_true", help="Only build result table, do not export skill files")
    xlsx.add_argument("--deleted-only", action="store_true", help="Only include skills deleted from current main branch")
    xlsx.add_argument("--report-json", type=Path, default=Path("output/export_from_xlsx_report.json"))

    deleted = sub.add_parser("export-deleted-all", help="Export all deleted skills from inventory (no input xlsx)")
    deleted.add_argument("--inventory-csv", type=Path, default=Path("output/inventory_all_skills.csv"))
    deleted.add_argument("--output-xlsx", type=Path, default=Path("output/clawhub_deleted_all.xlsx"))
    deleted.add_argument("--out-dir", type=Path, default=Path("output/recovered_deleted_all"))
    deleted.add_argument("--mode", choices=["latest", "all"], default="latest")
    deleted.add_argument("--force", action="store_true", help="Overwrite existing exported folders")
    deleted.add_argument("--skip-extract", action="store_true", help="Only build result table, do not export skill files")
    deleted.add_argument("--report-json", type=Path, default=Path("output/export_deleted_all_report.json"))

    full = sub.add_parser(
        "export-full-clawhub",
        help="Export the full unique ClawHub skill set aggregated from all historical commits into data/skills/clawhub",
    )
    full.add_argument(
        "--out-root",
        type=Path,
        default=Path("data/skills/clawhub"),
        help="Output root directory for the full ClawHub dataset",
    )
    full.add_argument("--force", action="store_true", help="Overwrite existing exported folders")
    full.add_argument("--skip-extract", action="store_true", help="Only build inventory/manifest, do not export skill files")

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    repo = args.repo.resolve()
    if args.command == "init-repo":
        init_repo(repo=repo, remote=args.remote, ref=args.ref, clone_mode=args.clone_mode)
        print(f"[ok] repo ready: {repo}")
        return

    if not (repo / ".git").exists():
        print(f"[info] repo not initialized, auto-initializing: {repo}")
        init_repo(repo=repo, remote=args.remote, ref=args.ref, clone_mode="full")
        print(f"[ok] repo ready: {repo}")

    if args.command == "build-inventory":
        payload = write_inventory(repo=repo, ref=args.ref, out_csv=args.out_csv, out_json=args.out_json)
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    if args.command == "export":
        report = export_targets(
            repo=repo,
            ref=args.ref,
            targets_file=args.targets,
            inventory_csv=args.inventory_csv,
            out_dir=args.out_dir,
            mode=args.mode,
            force=args.force,
        )
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return

    if args.command == "export-from-xlsx":
        report = export_from_xlsx(
            repo=repo,
            ref=args.ref,
            inventory_csv=args.inventory_csv,
            input_xlsx=args.input_xlsx,
            output_xlsx=args.output_xlsx,
            out_dir=args.out_dir,
            mode=args.mode,
            force=args.force,
            skip_extract=args.skip_extract,
            deleted_only=args.deleted_only,
            report_json=args.report_json,
        )
        print(
            json.dumps(
                {
                    "input_xlsx": report["input_xlsx"],
                    "output_xlsx": report["output_xlsx"],
                    "rows_total": report["rows_total"],
                    "rows_with_match": report["rows_with_match"],
                    "unique_matched_slugs": report["unique_matched_slugs"],
                    "rows_extracted_success": report["rows_extracted_success"],
                    "report_json": str(args.report_json),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    if args.command == "export-deleted-all":
        report = export_deleted_all(
            repo=repo,
            ref=args.ref,
            inventory_csv=args.inventory_csv,
            output_xlsx=args.output_xlsx,
            out_dir=args.out_dir,
            mode=args.mode,
            force=args.force,
            skip_extract=args.skip_extract,
            report_json=args.report_json,
        )
        print(
            json.dumps(
                {
                    "output_xlsx": report["output_xlsx"],
                    "deleted_slug_count": report["deleted_slug_count"],
                    "rows_written": report["rows_written"],
                    "rows_extracted_success": report["rows_extracted_success"],
                    "report_json": str(args.report_json),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    if args.command == "export-full-clawhub":
        report = export_full_clawhub_dataset(
            repo=repo,
            ref=args.ref,
            out_root=args.out_root,
            force=args.force,
            skip_extract=args.skip_extract,
        )
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return


if __name__ == "__main__":
    main()
